"""
Excel Processor with PII Detection and Redaction
Processes Excel files uploaded from React frontend, detects PII, and provides redacted versions
"""
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from pathlib import Path
from typing import Dict, Any, List, Tuple, Optional
from datetime import datetime
import re
import logging

from .base_processor import BaseFileProcessor
from app.core.pii_detector import PIIDetector

logger = logging.getLogger(__name__)

class ExcelProcessor(BaseFileProcessor):
    """
    Processes Excel files and performs PII detection and redaction
    Maintains Excel structure, formulas, and formatting while redacting sensitive data
    """
    
    def __init__(self, file_path: str | Path):
        super().__init__(file_path)
        
        # Validate that this is actually an Excel file
        if self.file_type != 'excel':
            raise ValueError(f"ExcelProcessor can only process Excel files, got: {self.file_type}")
        
        # Initialize PII detector
        self.pii_detector = PIIDetector()
        
        # Storage for analysis results
        self.workbook = None
        self.analysis_results = {}
        self.pii_findings = []
        
    def _load_workbook(self) -> openpyxl.Workbook:
        """Load Excel workbook with openpyxl for structure preservation"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)
            return self.workbook
        except Exception as e:
            raise ValueError(f"Failed to load Excel file: {str(e)}")
    
    def _analyze_worksheet_structure(self, worksheet) -> Dict[str, Any]:
        """
        Analyze worksheet structure to understand data layout
        
        Returns:
            Dict with structure analysis including headers, data ranges, etc.
        """
        structure = {
            'name': worksheet.title,
            'max_row': worksheet.max_row,
            'max_column': worksheet.max_column,
            'headers': [],
            'data_ranges': [],
            'merged_cells': list(worksheet.merged_cells.ranges),
            'has_formulas': False,
            'cell_types': {}
        }
        
        # Analyze first few rows to detect headers
        for row in range(1, min(6, worksheet.max_row + 1)):
            row_data = []
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                if cell.value is not None:
                    row_data.append(str(cell.value))
                    
                    # Check for formulas
                    if hasattr(cell, 'formula') and cell.formula:
                        structure['has_formulas'] = True
                        
            if row_data and row == 1:
                structure['headers'] = row_data
                
        return structure
    
    def _detect_pii_in_cell(self, cell_value: Any, cell_ref: str) -> List[Dict[str, Any]]:
        """
        Detect PII in a single cell value
        
        Args:
            cell_value: The cell's value
            cell_ref: Cell reference (e.g., 'A1')
            
        Returns:
            List of PII findings
        """
        if cell_value is None:
            return []
            
        # Convert to string for analysis
        text_value = str(cell_value).strip()
        if not text_value:
            return []
        
        # Use PII detector
        pii_results = self.pii_detector.detect_pii(text_value)
        
        # Add cell reference to each finding
        for finding in pii_results:
            finding['cell_ref'] = cell_ref
            finding['original_value'] = cell_value
            
        return pii_results
    
    def _redact_cell_value(self, original_value: Any, pii_type: str, entity_id: str = None) -> str:
        """
        Redact a cell value based on PII type
        
        Args:
            original_value: Original cell value
            pii_type: Type of PII detected
            entity_id: Unique identifier for consistent redaction
            
        Returns:
            Redacted value
        """
        redaction_map = {
            'PERSON': '[NAME_REDACTED]',
            'EMAIL_ADDRESS': '[EMAIL_REDACTED]',
            'PHONE_NUMBER': '[PHONE_REDACTED]',
            'CREDIT_CARD': '[CARD_REDACTED]',
            'SSN': '[SSN_REDACTED]',
            'ORGANIZATION': '[ORG_REDACTED]',
            'LOCATION': '[LOCATION_REDACTED]',
            'DATE_TIME': '[DATE_REDACTED]',
            'IBAN_CODE': '[IBAN_REDACTED]',
            'IP_ADDRESS': '[IP_REDACTED]',
            'URL': '[URL_REDACTED]'
        }
        
        # Use entity ID for consistent redaction if provided
        if entity_id:
            return f"[{pii_type}_{entity_id}]"
            
        return redaction_map.get(pii_type, '[PII_REDACTED]')
    
    def extract_text(self) -> Dict[str, Any]:
        """
        Extract text and analyze PII in Excel file
        
        Returns:
            Dict containing extracted text, PII findings, and metadata
        """
        try:
            # Load workbook
            workbook = self._load_workbook()
            
            all_text_content = []
            total_pii_findings = []
            worksheets_analysis = {}
            
            # Process each worksheet
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                # Analyze structure
                structure = self._analyze_worksheet_structure(worksheet)
                worksheets_analysis[sheet_name] = structure
                
                # Extract text and detect PII
                sheet_text = []
                sheet_pii_findings = []
                
                for row in range(1, worksheet.max_row + 1):
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell_ref = f"{sheet_name}!{cell.coordinate}"
                        
                        if cell.value is not None:
                            cell_text = str(cell.value)
                            sheet_text.append(f"{cell_ref}: {cell_text}")
                            
                            # Detect PII in this cell
                            pii_findings = self._detect_pii_in_cell(cell.value, cell_ref)
                            sheet_pii_findings.extend(pii_findings)
                
                all_text_content.extend(sheet_text)
                total_pii_findings.extend(sheet_pii_findings)
            
            # Store findings for redaction
            self.pii_findings = total_pii_findings
            self.analysis_results = worksheets_analysis
            
            # Prepare summary
            pii_summary = {}
            for finding in total_pii_findings:
                pii_type = finding['entity_type']
                if pii_type not in pii_summary:
                    pii_summary[pii_type] = 0
                pii_summary[pii_type] += 1
            
            # Calculate processing confidence based on successful cell analysis
            total_cells = sum(ws['max_row'] * ws['max_column'] for ws in worksheets_analysis.values())
            processed_cells = sum(1 for _ in all_text_content if _)  # Count non-empty cells
            
            # Calculate confidence: higher when more cells processed successfully
            processing_confidence = min(95, max(75, int((processed_cells / max(total_cells, 1)) * 100))) if total_cells > 0 else 85
            
            return {
                'text': '\n'.join(all_text_content[:100]),  # Limit for display
                'confidence': processing_confidence,  # Add confidence metric
                'pii_findings': total_pii_findings,
                'pii_summary': pii_summary,
                'total_pii_count': len(total_pii_findings),
                'worksheets': list(worksheets_analysis.keys()),
                'total_cells_processed': sum(ws['max_row'] * ws['max_column'] for ws in worksheets_analysis.values()),
                'metadata': {
                    'worksheets_count': len(workbook.sheetnames),
                    'worksheets_analysis': worksheets_analysis
                }
            }
            
        except Exception as e:
            logger.error(f"Error processing Excel file {self.file_path}: {str(e)}")
            return {
                'text': f'Error processing Excel file: {str(e)}',
                'pii_findings': [],
                'pii_summary': {},
                'total_pii_count': 0,
                'metadata': {'error': str(e)}
            }
    
    def create_redacted_excel(self, output_path: Path = None, original_filename: str = None) -> Path:
        """
        Create a redacted version of the Excel file
        
        Args:
            output_path: Where to save the redacted file
            original_filename: Original filename for better naming
            
        Returns:
            Path to the redacted Excel file
        """
        if not self.workbook:
            self._load_workbook()
            
        if not output_path:
            # Save to processed_files directory for download
            processed_dir = Path("data/processed_files")
            processed_dir.mkdir(parents=True, exist_ok=True)
            
            # Use original filename if provided, otherwise use current file stem
            if original_filename:
                # Extract clean name from original filename
                clean_name = Path(original_filename).stem
            else:
                # Fallback to current file stem
                original_name = self.file_path.stem
                # Remove any existing timestamp or hash suffixes
                clean_name = original_name.split('_')[0] if '_' in original_name else original_name
            
            # Create readable redacted filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_path = processed_dir / f"{clean_name}_REDACTED_{timestamp}.xlsx"
        
        # Create a copy of the workbook for redaction
        redacted_workbook = openpyxl.load_workbook(self.file_path)
        
        # Apply redactions
        entity_counter = {}  # For consistent entity IDs
        
        for finding in self.pii_findings:
            cell_ref = finding['cell_ref']
            sheet_name, cell_coord = cell_ref.split('!')
            
            worksheet = redacted_workbook[sheet_name]
            cell = worksheet[cell_coord]
            
            # Generate consistent entity ID
            pii_type = finding['entity_type']
            if pii_type not in entity_counter:
                entity_counter[pii_type] = 1
            else:
                entity_counter[pii_type] += 1
            
            entity_id = entity_counter[pii_type]
            
            # Redact the cell
            redacted_value = self._redact_cell_value(
                finding['original_value'], 
                pii_type, 
                entity_id
            )
            
            cell.value = redacted_value
            
            # Highlight redacted cells (optional)
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Save redacted workbook
        redacted_workbook.save(output_path)
        redacted_workbook.close()
        
        return output_path